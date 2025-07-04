import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from pathlib import Path

def parse_json_files(folder_path, sheet, row_start=1):
    # Определяем возможные поля
    fields = ["id", "jp", "en", "ru", "en_voice", "jp_voice"]
    
    # Определяем желтую заливку
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Рекурсивно обходим все файлы в папке
    for item in Path(folder_path).rglob("*.json"):
        with open(item, 'r', encoding='utf-8') as f:
            try:
                data = json.load(f)
                for entry in data:
                    # Записываем каждую группу
                    for field in fields:
                        sheet[f"A{row_start}"] = field
                        sheet[f"B{row_start}"] = entry.get(field, "")
                        # Применяем желтую заливку для "en_voice" если значение в B пустое
                        if field == "en_voice" and not entry.get(field):
                            sheet[f"A{row_start}"].fill = yellow_fill
                        row_start += 1
                    # Добавляем пустую строку между группами
                    row_start += 1
            except json.JSONDecodeError:
                print(f"Ошибка при чтении файла: {item}")
    return row_start

def main():
    # Путь к папке с JSON-файлами
    base_path = "nier_text_json"
    
    # Создаем новый Excel-файл
    wb = Workbook()
    
    # Удаляем дефолтный лист
    wb.remove(wb.active)
    
    # Обрабатываем каждую папку (core, ph1, ph2 и т.д.)
    for folder in Path(base_path).iterdir():
        if folder.is_dir():
            # Создаем новый лист с именем папки
            sheet = wb.create_sheet(title=folder.name)
            
            # Парсим JSON-файлы и записываем в лист
            parse_json_files(folder, sheet)
    
    # Сохраняем Excel-файл
    wb.save("nier_subtitles.xlsx")

if __name__ == "__main__":
    main()